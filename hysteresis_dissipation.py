# -*- coding: ascii -*-
"""
Is the model dissipating energy before the mechanism forms?

Builds base-shear vs top-quarter-displacement hysteresis loops for the model and,
where available, for US-1, and measures the enclosed area. The question this
answers: in the pre-mechanism runs (roughly 5-13), is the model behaving
near-elastically -- thin, closed loops -- while the specimen was already
dissipating? If so, G_I / G_II are too brittle and damage localises suddenly at
run 14 instead of distributing, which is the central discrepancy in
MODELLING_DISCREPANCIES.md section 3.

DEFINITIONS (following the source paper, Structures 66 (2024) 106815)

  U = average of the two top-quarter potentiometers, at Z = 2.06 m   [Eq. 1]
      model      : 0.5*(Channel_3 + Channel_4) - Channel_5   (table-referenced,
                   because the experimental sensors are frame-mounted)
      experiment : U_avg from processed_globalzero (already relative)
  F = base shear                                                     [Eq. 2]
      model      : total_shear  (cstav + joist shears; the FISH comment states
                   this is the quantity comparable to the experimental
                   V = sum(m*a)). Use --force cstav to compare against the
                   contact base shear alone.
      experiment : base_shear_kN from processed_globalzero, which is
                   sum(a_j * m_j) from the accelerometers

  Caveat worth remembering: the experimental F is an INERTIAL base shear derived
  from accelerometers, the model's is a CONTACT force. They agree for a wall in
  dynamic equilibrium but diverge during impacts, so treat small differences in
  loop area with corresponding caution.

METRICS (per run, and per half-cycle)

  E_diss    CUMULATIVE hysteretic energy: the loop is split into half-cycles at
            zero crossings of U and the absolute enclosed area of each is summed,
            in J (1 kN * 1 mm = 1 J). Summing absolute areas is the standard
            measure from cyclic testing and is necessary here: a whole-run signed
            area lets sub-loops of opposite handedness cancel, which made US-1
            run 18 report 4 J when its displacement range was -9.0 to +2.1 mm.
  E_net     the whole-run signed area, kept for reference. Where E_net is much
            smaller than E_diss, the response contains counter-rotating sub-loops.
  fullness  energy-weighted mean of the per-half-cycle shape factor
            area_i / (4 * F_peak_i * U_peak_i). Each cycle is normalised by its
            OWN peaks -- normalising a cumulative sum by the run's single peak
            lets many small cycles drive the value above 1, which is impossible.
            0.0 = a straight line (perfectly elastic), 1.0 = a rectangular
            elastoplastic loop. This is the number that answers the question.
  xi_eq     equivalent viscous damping, 2/pi * fullness

Usage
    python hysteresis_dissipation.py [SIM_DIR] [--runs 5 12] [--test 9]
                                     [--force total_shear|cstav] [--no-exp]
                                     [--out-prefix NAME]

SIM_DIR defaults to safego_paths.sim_dir(). Writes
<SIM>/postproc/<prefix>_loops.png and <prefix>_dissipation.csv.
"""
import argparse, csv, glob, math, os, sys
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

try:
    import safego_paths as sp
    _HAVE_SP = True
except ImportError:
    _HAVE_SP = False

H_TOP = 2.06

# Repo figure convention (CLAUDE.md section 5): red = simulation, blue = US-1.
C_SIM = "tab:red"
C_EXP = "tab:blue"


# ---------------------------------------------------------------- io helpers
def read_hist(path):
    """3DEC history export: 2 header lines, then whitespace-separated time/value."""
    d = np.genfromtxt(path, skip_header=2)
    m = np.isfinite(d[:, 0]) & np.isfinite(d[:, 1])
    return d[m, 0], d[m, 1]


def find_channel(folder, pattern):
    hits = glob.glob(os.path.join(folder, "*" + pattern + "*"))
    return hits[0] if hits else None


def load_model_run(folder, force_key):
    """Return (U_mm, F_kN) for one model run folder, or None."""
    def ch(pat):
        p = find_channel(folder, pat)
        return read_hist(p) if p else (None, None)

    _, c3 = ch("Channel_3_")
    _, c4 = ch("Channel_4_")
    _, c5 = ch("Channel_5_DispTable")
    _, f = ch(force_key)
    if c3 is None or c4 is None or c5 is None or f is None:
        return None
    n = min(len(c3), len(c4), len(c5), len(f))
    t, _ = ch("Channel_3_")
    dt = float(np.median(np.diff(t[:n]))) if n > 2 else 0.0
    U = (0.5 * (c3[:n] + c4[:n]) - c5[:n]) * 1000.0        # mm, table-referenced
    F = f[:n]                                              # kN
    return U - U[0], F - F[0], dt


def load_exp_run(proc_dir, test_no, run_no):
    """Return (U_mm, F_kN) from processed_globalzero, or None."""
    try:
        import openpyxl
    except ImportError:
        return None
    p = os.path.join(str(proc_dir),
                     "Test{}Run{}_processed_globalzero.xlsx".format(test_no, run_no))
    if not os.path.isfile(p):
        return None
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows)                                             # header
    U, F = [], []
    for r in rows:
        if r[1] is None or r[3] is None:
            continue
        U.append(float(r[1]) * 1000.0)                     # U_avg m -> mm
        F.append(float(r[3]))                              # base_shear_kN
    wb.close()
    if not U:
        return None
    U = np.asarray(U); F = np.asarray(F)
    return U - U[0], F - F[0], 0.005    # processed exports are 5 ms


# ------------------------------------------------------------------- metrics
def lowpass(x, dt, fc):
    """Zero-phase low-pass. The model's contact forces carry high-frequency
    chatter from rocking impacts that no accelerometer-derived experimental base
    shear contains; comparing raw loops is not like-for-like. Metrics are
    reported both raw and filtered so the sensitivity is visible."""
    if fc is None or fc <= 0 or dt <= 0:
        return x
    fs = 1.0 / dt
    if fc >= 0.5 * fs:
        return x
    try:
        from scipy.signal import butter, filtfilt
        b, a = butter(4, fc / (0.5 * fs), btype="low")
        return filtfilt(b, a, x)
    except Exception:
        w = max(1, int(round(fs / fc)))
        return np.convolve(x, np.ones(w) / w, mode="same")


def shoelace_area(x, y):
    """Signed area enclosed by the (possibly self-intersecting) path, closed
    back to its start. For a hysteresis loop traversed consistently this is the
    dissipated energy; the sign carries the traversal direction."""
    x = np.asarray(x, float); y = np.asarray(y, float)
    if len(x) < 3:
        return 0.0
    xs = np.append(x, x[0]); ys = np.append(y, y[0])
    return 0.5 * float(np.sum(xs[:-1] * ys[1:] - xs[1:] * ys[:-1]))


def split_cycles(U, min_amp_frac=0.15):
    """Segment the history into half-cycles at zero crossings of U, keeping only
    excursions that reach at least min_amp_frac of the run's peak. Returns a list
    of (start, stop) index pairs."""
    if len(U) < 8:
        return []
    thr = min_amp_frac * np.max(np.abs(U))
    if thr <= 0:
        return []
    sign = np.sign(U)
    idx = np.where(np.diff(sign) != 0)[0]
    segs = []
    for a, b in zip(np.r_[0, idx], np.r_[idx, len(U) - 1]):
        if b - a < 4:
            continue
        if np.max(np.abs(U[a:b])) >= thr:
            segs.append((int(a), int(b)))
    return segs


def run_metrics(U, F, dt=0.0, fc=None):
    """Whole-run and per-cycle dissipation metrics. If fc is given, both signals
    are low-passed first and the raw values are also reported."""
    raw = None
    if fc:
        _rs = split_cycles(U)
        _c = [abs(shoelace_area(U[a:b], F[a:b])) for a, b in _rs]
        _r = [4.0 * float(np.max(np.abs(F[a:b]))) * float(np.max(np.abs(U[a:b])))
              for a, b in _rs]
        raw_E = float(np.sum(_c)) if _c else abs(shoelace_area(U, F))
        _d = float(np.sum(_r)) if _r else 0.0
        raw = (raw_E, (raw_E / _d) if _d > 0 else float("nan"))
        U = lowpass(U, dt, fc); F = lowpass(F, dt, fc)
    E_net = abs(shoelace_area(U, F))                       # whole-run signed area
    segs = split_cycles(U)
    Up = float(np.max(np.abs(U))); Fp = float(np.max(np.abs(F)))

    # Per-half-cycle areas, each normalised by ITS OWN peaks. Normalising a
    # cumulative sum by the run's single peak is wrong -- it lets many small
    # cycles push "fullness" above 1, which is impossible for a real loop.
    cyc, ref = [], []
    for a, b in segs:
        u, f = U[a:b], F[a:b]
        cyc.append(abs(shoelace_area(u, f)))
        ref.append(4.0 * float(np.max(np.abs(f))) * float(np.max(np.abs(u))))
    E = float(np.sum(cyc)) if cyc else E_net               # cumulative energy, J
    denom = float(np.sum(ref)) if ref else (4.0 * Fp * Up)
    # energy-weighted mean of the per-cycle shape factor; bounded by 1 for a
    # physical loop, 0 for a straight line
    full = (E / denom) if denom > 0 else float("nan")
    out = {
        "U_peak_mm": Up,
        "F_peak_kN": Fp,
        "E_diss_J": E,
        "E_net_J": E_net,
        "fullness": full,
        "xi_eq": full * 2.0 / math.pi,
    }
    out["n_cycles"] = len(segs)
    out["E_diss_largest_cycle_J"] = max(cyc) if cyc else float("nan")
    # secant stiffness over the peak excursion, kN/mm
    i = int(np.argmax(np.abs(U)))
    out["k_secant_kN_per_mm"] = (abs(F[i]) / abs(U[i])) if abs(U[i]) > 1e-9 else float("nan")
    if raw is not None:
        out["E_diss_J_raw"] = raw[0]
        out["fullness_raw"] = raw[1]
    return out


# ---------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    default_sim = str(sp.sim_dir()) if _HAVE_SP else "stratC_results_NODAMP_v6_NEW"
    ap.add_argument("sim_dir", nargs="?", default=default_sim)
    ap.add_argument("--runs", nargs=2, type=int, default=[5, 12],
                    help="inclusive run range (default 5 12, the pre-mechanism band)")
    ap.add_argument("--test", type=int, default=9, choices=(9, 12))
    ap.add_argument("--force", default="total_shear",
                    choices=("total_shear", "cstav"),
                    help="model force channel (default total_shear)")
    ap.add_argument("--no-exp", action="store_true", help="skip the experimental overlay")
    ap.add_argument("--lowpass", type=float, default=50.0,
                    help="low-pass both signals at this Hz before measuring and "
                         "plotting (default 50; 0 disables). The model's contact "
                         "forces carry impact chatter the experimental inertial "
                         "base shear does not.")
    ap.add_argument("--out-prefix", default="hyst")
    args = ap.parse_args()

    force_key = "total_shear" if args.force == "total_shear" else "cstav_BaseShear"
    sim = args.sim_dir
    pp = os.path.join(sim, "postproc")
    if not os.path.isdir(pp):
        os.makedirs(pp)
    proc_dir = (sp.processed_dir() if _HAVE_SP
                else os.path.join("EXP_DATA", "processed_globalzero"))

    lo, hi = args.runs
    wanted = list(range(lo, hi + 1))
    folders = {}
    for f in sorted(glob.glob(os.path.join(sim, "Run*"))):
        if not os.path.isdir(f):
            continue
        try:
            folders[int(os.path.basename(f)[3:5])] = f
        except ValueError:
            pass

    rows = []
    print("model force channel : {}".format(force_key))
    print("displacement        : 0.5*(Ch3+Ch4) - Ch5 at Z=2.06 m  (paper Eq. 1)")
    print("")
    print("{:>4} | {:>9} {:>9} {:>10} {:>9} {:>7} | {:>9} {:>9} {:>10} {:>9} {:>7}".format(
        "run", "U_pk mm", "F_pk kN", "E_diss J", "fullness", "xi_eq",
        "U_pk mm", "F_pk kN", "E_diss J", "fullness", "xi_eq"))
    print("{:>4} | {:^49} | {:^49}".format("", "MODEL", "US-{}".format(1 if args.test == 9 else 2)))

    data = {}
    for rn in wanted:
        row = {"run": rn}
        m = load_model_run(folders[rn], force_key) if rn in folders else None
        e = None if args.no_exp else load_exp_run(proc_dir, args.test, rn)
        data[rn] = (m, e)
        cells = []
        for tag, d in (("sim", m), ("exp", e)):
            if d is None:
                cells.append(("-",) * 5)
                continue
            mt = run_metrics(d[0], d[1], d[2], args.lowpass)
            for k, v in mt.items():
                row["{}_{}".format(tag, k)] = v
            cells.append(("{:9.3f}".format(mt["U_peak_mm"]),
                          "{:9.3f}".format(mt["F_peak_kN"]),
                          "{:10.2f}".format(mt["E_diss_J"]),
                          "{:9.4f}".format(mt["fullness"]),
                          "{:7.4f}".format(mt["xi_eq"])))
        print("{:>4} | {} {} {} {} {} | {} {} {} {} {}".format(rn, *(cells[0] + cells[1])))
        rows.append(row)

    # ---- verdict -----------------------------------------------------------
    sf = [r["sim_fullness"] for r in rows if r.get("sim_fullness") == r.get("sim_fullness")]
    ef = [r["exp_fullness"] for r in rows if r.get("exp_fullness") == r.get("exp_fullness")]
    print("\n" + "=" * 78)
    if sf:
        print("MODEL   mean loop fullness over runs {}-{} : {:.4f}".format(lo, hi, np.mean(sf)))
    if ef:
        print("US-{}    mean loop fullness over runs {}-{} : {:.4f}".format(
            1 if args.test == 9 else 2, lo, hi, np.mean(ef)))
    if sf and ef and np.mean(ef) > 0:
        print("ratio model/experiment                    : {:.2f}x".format(np.mean(sf) / np.mean(ef)))
    print("""
Reading the number: fullness is the loop area normalised by 4*F_peak*U_peak.
  < 0.05   essentially elastic - the model is storing energy, not dissipating it
  0.1-0.3  pinched rocking hysteresis, which is what the paper describes
  > 0.4    fat, elastoplastic loops
If the model sits far below the experiment here, G_I/G_II are too brittle: nothing
dissipates until the joints fail outright, and then a mechanism forms all at once.""")
    print("=" * 78)

    # ---- csv ---------------------------------------------------------------
    fields = ["run"]
    for r in rows:
        for k in r:
            if k not in fields:
                fields.append(k)
    out_csv = os.path.join(pp, "{}_dissipation.csv".format(args.out_prefix))
    with open(out_csv, "w") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow({k: (round(v, 6) if isinstance(v, float) else v)
                        for k, v in r.items()})
    print("\n-> {}".format(out_csv))

    # ---- figure ------------------------------------------------------------
    n = len(wanted)
    ncol = 4
    nrow = int(math.ceil(n / float(ncol)))
    fig, axes = plt.subplots(nrow, ncol, figsize=(4.0 * ncol, 3.4 * nrow),
                             squeeze=False)
    for ax, rn in zip(axes.ravel(), wanted):
        m, e = data[rn]
        if e is not None:
            ax.plot(lowpass(e[0], e[2], args.lowpass), lowpass(e[1], e[2], args.lowpass),
                    color=C_EXP, lw=0.7, alpha=0.85,
                    label="US-{}".format(1 if args.test == 9 else 2))
        if m is not None:
            ax.plot(lowpass(m[0], m[2], args.lowpass), lowpass(m[1], m[2], args.lowpass),
                    color=C_SIM, lw=0.7, alpha=0.85, label="model")
        ax.axhline(0, color="k", lw=0.5); ax.axvline(0, color="k", lw=0.5)
        ax.set_title("Run {:02d}".format(rn), fontsize=10)
        ax.grid(alpha=0.15)
        txt = []
        if m is not None:
            mm = run_metrics(m[0], m[1], m[2], args.lowpass); txt.append("sim  E={:.1f} J  full={:.3f}".format(
                mm["E_diss_J"], mm["fullness"]))
        if e is not None:
            me = run_metrics(e[0], e[1], e[2], args.lowpass); txt.append("exp  E={:.1f} J  full={:.3f}".format(
                me["E_diss_J"], me["fullness"]))
        if txt:
            ax.text(0.03, 0.97, "\n".join(txt), transform=ax.transAxes,
                    va="top", ha="left", fontsize=7.5,
                    bbox=dict(fc="white", ec="none", alpha=0.75))
    for ax in axes.ravel()[n:]:
        ax.axis("off")
    for ax in axes[-1]:
        ax.set_xlabel("Top quarter OOP displacement (mm)")
    for ax in axes[:, 0]:
        ax.set_ylabel("Base shear (kN)")
    h, l = axes[0, 0].get_legend_handles_labels()
    if h:
        fig.legend(h, l, loc="upper right", fontsize=9)
    fig.suptitle("Base shear vs top-quarter OOP displacement, runs {}-{}"
                 "   (model force = {}, low-pass {:.0f} Hz)".format(lo, hi, force_key, args.lowpass),
                 fontsize=12, fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.96])
    out_png = os.path.join(pp, "{}_loops.png".format(args.out_prefix))
    fig.savefig(out_png, dpi=200, bbox_inches="tight")
    print("-> {}".format(out_png))


if __name__ == "__main__":
    main()
