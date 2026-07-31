# -*- coding: ascii -*-
"""
figP8: max and normalized horizontal displacement profiles over height
       (Moshfeghi-style, 4 rows: US-1, US-2, NUM no-damp, NUM ratcheting).
figP9: base-shear vs top-displacement hysteresis, NUM vs US-1
       (processed_globalzero U_avg / base_shear_kN), selected runs.

Profiles use wall-relative displacement (exp sensors are table-mounted;
sim channels are referenced to Channel_5). Envelope = min/max of
(x - x_run_start) over all runs. Normalized panels: first and last run
peak-abs profile, normalized by the 2.06 m sensor value.

Usage (staged; state in /tmp/prof_state.json):
  python profile_hysteresis.py exp9  EXP_DIR LO HI
  python profile_hysteresis.py exp12 EXP_DIR LO HI
  python profile_hysteresis.py sim
  python profile_hysteresis.py simratch
  python profile_hysteresis.py fig
"""
import sys, json, glob
import numpy as np
from pathlib import Path

HERE = Path(__file__).resolve().parent
SIM = HERE / "stratC_results_NODAMP_v6_NEW"
SIM_RATCH = HERE / "stratC_results_RATCHETING"
PP = SIM / "postproc"
PROC9 = Path("/sessions/epic-stoic-galileo/mnt/Hanze--EXP_DATA/processed_globalzero")
STATE = Path("/tmp/prof_state.json")

# heights: (label, z) per case; exp top pin at 2.85 m added in the figure
H_EXP9 = [("bot", 0.66), ("mid", 1.26), ("top", 2.06)]
H_EXP12 = [("bot", 0.60), ("mid", 1.26), ("top", 2.06)]
H_SIM = [("bot", 0.66), ("mid", 1.26), ("top", 2.06), ("beam", 2.68)]
HYST_RUNS = [7, 15, 21, 24]


def load_state():
    return json.loads(STATE.read_text()) if STATE.exists() else {}


def save_state(st):
    STATE.write_text(json.dumps(st))


def read_hist(fpath):
    d = np.genfromtxt(str(fpath), skip_header=2)
    m = np.isfinite(d[:, 0]) & np.isfinite(d[:, 1])
    return d[m, 0], d[m, 1]


def do_exp(st, key, exp_dir, prefix, lo, hi):
    import openpyxl
    st.setdefault(key, {})
    for rn in range(lo, hi + 1):
        f = Path(exp_dir) / ("%s%d.xlsx" % (prefix, rn))
        if not f.exists():
            print("missing:", f.name); continue
        wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        next(rows)
        c1, c2, c4 = [], [], []
        for r in rows:
            if r[1] is None:
                continue
            c1.append(r[2]); c2.append(r[3]); c4.append(r[5])
        wb.close()
        rec = {}
        for name, v in (("bot", c1), ("mid", c2), ("top", c4)):
            x = np.asarray(v, float) * 1000.0
            x = x - x[0]
            rec[name] = [round(float(x.min()), 3), round(float(x.max()), 3)]
        st[key][str(rn)] = rec
        print("%s run %2d: top %7.2f..%7.2f mm" % (key, rn,
              rec["top"][0], rec["top"][1]))
        save_state(st)


def do_sim(st, sim_dir, key, lo=1, hi=25):
    st.setdefault(key, {})
    chan = {"bot": "*Channel_1_DispBot*", "mid": "*Channel_2_DispMid*",
            "top": "*Channel_4_DispTopQRight*", "beam": "*Channel_20_TopBeam*"}
    for folder in sorted(Path(sim_dir).iterdir()):
        if not (folder.is_dir() and folder.name.startswith("Run")):
            continue
        rn = int(folder.name[3:5])
        if not (lo <= rn <= hi) or str(rn) in st[key]:
            continue
        ft = glob.glob(str(folder / "*Channel_5_DispTable*"))
        if not ft:
            continue
        _, tab = read_hist(ft[0])
        rec = {}
        for name, pat in chan.items():
            fl = glob.glob(str(folder / pat))
            if not fl:
                continue
            _, x = read_hist(fl[0])
            n = min(len(x), len(tab))
            rel = (x[:n] - tab[:n]) * 1000.0
            rel = rel - rel[0]
            rec[name] = [round(float(rel.min()), 3), round(float(rel.max()), 3)]
        st[key][str(rn)] = rec
        print("%s run %2d: top %7.2f..%7.2f mm" % (key, rn,
              rec["top"][0], rec["top"][1]))
    save_state(st)


def overlay_fig(plt, st):
    """figP8b: EXP and NUM displacement profiles overlaid.
    Colour = family (blue EXP, red NUM); linestyle/marker = case."""
    cases = [("US-1 (Test 9)", "exp9", H_EXP9, True,
              "royalblue", "-", "o", True),
             ("US-2 (Test 12)", "exp12", H_EXP12, True,
              "royalblue", "--", "^", True),
             ("NUM: undamped", "simprof", H_SIM, False,
              "red", "-", "s", False),
             ("NUM: 3% Rayleigh", "simratchprof", H_SIM, False,
              "purple", "--", "D", False)]

    def case_data(key, hs, is_exp):
        d = st.get(key, {})
        runs = sorted(int(k) for k in d)
        env_runs = ([r for r in runs
                     if max(abs(v) for v in d[str(r)]["top"]) <= 35.0]
                    if not is_exp else runs)
        return d, runs, env_runs

    fig, (axL, axM, axR) = plt.subplots(
        1, 3, figsize=(16.5, 6.6), sharey=True,
        gridspec_kw={"width_ratios": [1.45, 1.0, 1.0]})

    for ttl, key, hs, is_exp, colr, ls, mk, filled in cases:
        d, runs, env_runs = case_data(key, hs, is_exp)
        if not runs:
            continue
        mfc = colr if filled else "white"
        kw = dict(color=colr, ls=ls, marker=mk, ms=8, mfc=mfc, mec=colr,
                  mew=1.5, lw=2.0)

        # envelope
        zs = [0.0] + [z for _, z in hs]
        env_min = [0.0] + [min(d[str(r)][n][0] for r in env_runs)
                           for n, _ in hs]
        env_max = [0.0] + [max(d[str(r)][n][1] for r in env_runs)
                           for n, _ in hs]
        if is_exp:
            zs, env_min, env_max = (zs + [2.85], env_min + [0.0],
                                    env_max + [0.0])
        axL.plot(env_min, zs, label=ttl, **kw)
        axL.plot(env_max, zs, **kw)

        # normalized profiles, first and last run
        for ax, rn in ((axM, runs[0]),
                       (axR, env_runs[-1] if not is_exp else runs[-1])):
            prof = [max(abs(v) for v in d[str(rn)][n]) for n, _ in hs]
            ref = prof[[n for n, _ in hs].index("top")] or 1.0
            xs = [0.0] + [pv / ref for pv in prof]
            zz = [0.0] + [z for _, z in hs]
            if is_exp:
                xs, zz = xs + [0.0], zz + [2.85]
            ax.plot(xs, zz, **kw)

    axL.axvline(0, color="0.55", lw=0.8)
    axL.set_xlim(-35, 35)
    axL.set_ylim(0, 3)
    axL.set_xlabel("Max. horizontal displacement (mm)", fontsize=15)
    axL.set_ylabel("Height (m)", fontsize=15)
    axL.set_title("Whole-test envelope", fontsize=14)
    axL.legend(fontsize=11.5, loc="lower left", handlelength=2.6)
    axM.set_title("Normalized profile, first run", fontsize=14)
    axR.set_title("Normalized profile, last run", fontsize=14)
    for ax in (axM, axR):
        ax.set_xlim(-0.05, 1.35)
        ax.set_xlabel("Normalized displacement", fontsize=15)
    for ax in (axL, axM, axR):
        ax.tick_params(labelsize=13)
    fig.text(0.995, 0.01,
             "EXP: blue, filled markers.  NUM: red/purple, open markers. "
             "NUM collapse runs excluded from envelope and last run.",
             fontsize=9.5, color="0.35", ha="right")
    fig.tight_layout(rect=[0, 0.025, 1, 1])
    fig.savefig(str(PP / "figP8b_profiles_overlay.png"), dpi=300,
                bbox_inches="tight")
    plt.close(fig); print("-> figP8b_profiles_overlay.png")


RAW9 = Path("/sessions/epic-stoic-galileo/mnt/EXP_DATA")


def ch3_hysteresis_fig(plt):
    """figP10: base shear vs Channel 3 (top quarter left), US-1 vs NUM.
    Exp: ch3 from the raw xlsx paired row-wise with base_shear_kN from
    processed_globalzero (verified equal row counts). Both re-zeroed to
    the run start; sim ch3 is table-referenced."""
    import openpyxl
    fig, axes = plt.subplots(2, 2, figsize=(12.5, 9.5))
    for ax, rn in zip(axes.ravel(), HYST_RUNS):
        fr = RAW9 / ("Test9Run%d.xlsx" % rn)
        fp = PROC9 / ("Test9Run%d_processed_globalzero.xlsx" % rn)
        if fr.exists() and fp.exists():
            wb = openpyxl.load_workbook(str(fr), read_only=True,
                                        data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            next(rows)
            c3 = [r[4] for r in rows if r[1] is not None]
            wb.close()
            wb = openpyxl.load_workbook(str(fp), read_only=True,
                                        data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            next(rows)
            bs = [r[3] for r in rows if r[0] is not None]
            wb.close()
            n = min(len(c3), len(bs))
            x = np.asarray(c3[:n], float) * 1000.0
            ax.plot(x - x[0], np.asarray(bs[:n], float), "-",
                    color="royalblue", lw=0.7, alpha=0.85,
                    label="US-1 (Ch 3, base shear)")
        for sdir, colr, lbl in ((SIM_RATCH, "purple", "NUM 3% Rayleigh"),):
            folder = glob.glob(str(sdir / ("Run%02d_*" % rn)))
            if not folder:
                continue
            f3 = glob.glob(folder[0] + "/*Channel_3_DispTopQLeft*")
            f5 = glob.glob(folder[0] + "/*Channel_5_DispTable*")
            fbs = glob.glob(folder[0] + "/*cstav*")
            if f3 and f5 and fbs:
                _, x3 = read_hist(f3[0])
                _, x5 = read_hist(f5[0])
                _, v = read_hist(fbs[0])
                n = min(len(x3), len(x5), len(v))
                u = (x3[:n] - x5[:n]) * 1000.0
                ax.plot(u - u[0], v[:n], "-", color=colr, lw=0.8,
                        alpha=0.9, label=lbl + " (Ch 3, base-joint shear)")
        ax.axhline(0, color="0.4", lw=0.6)
        ax.axvline(0, color="0.4", lw=0.6)
        ax.set_title("Run %d" % rn, fontsize=12, fontweight="bold")
        ax.legend(fontsize=9, loc="upper right")
    for ax in axes[1]:
        ax.set_xlabel("Ch 3 OOP displacement (mm)", fontsize=12)
    for ax in axes[:, 0]:
        ax.set_ylabel("Base shear (kN)", fontsize=12)
    fig.tight_layout()
    fig.savefig(str(PP / "figP10_ch3_hysteresis.png"), dpi=300,
                bbox_inches="tight")
    plt.close(fig); print("-> figP10_ch3_hysteresis.png")


# ------------------------------------------------------------------ figures
def style():
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    plt.rcParams.update({
        "font.family": "serif", "font.serif": ["DejaVu Serif"],
        "mathtext.fontset": "dejavuserif", "axes.linewidth": 1.1,
        "xtick.direction": "in", "ytick.direction": "in",
        "xtick.top": True, "ytick.right": True,
        "axes.grid": True, "grid.linestyle": ":", "grid.alpha": 0.55,
        "legend.frameon": True, "legend.framealpha": 1.0,
        "legend.edgecolor": "0.7"})
    return plt


def profile_fig(plt, st):
    cases = [("US-1 (Test 9)", "exp9", H_EXP9, True),
             ("US-2 (Test 12)", "exp12", H_EXP12, True),
             ("NUM: no viscous damping", "simprof", H_SIM, False),
             ("NUM: 3% Rayleigh, asym. pulse (v5)", "simratchprof",
              H_SIM, False)]
    fig, axes = plt.subplots(4, 2, figsize=(11.5, 16.0))
    for i, (ttl, key, hs, is_exp) in enumerate(cases):
        d = st.get(key, {})
        if not d:
            continue
        runs = sorted(int(k) for k in d)
        first, last = runs[0], runs[-1]
        axL, axR = axes[i]

        # ---- left: whole-test envelope. For NUM cases the final collapse
        # run (top peak beyond the +-35 mm axis) is excluded and noted.
        env_runs = list(runs)
        note = ""
        if not is_exp:
            keep = [r for r in env_runs
                    if max(abs(v) for v in d[str(r)]["top"]) <= 35.0]
            if len(keep) < len(env_runs):
                excl = sorted(set(env_runs) - set(keep))
                note = "runs %s (collapse) excluded" % (
                    ", ".join(str(r) for r in excl))
                env_runs = keep
        zs = [0.0] + [z for _, z in hs]
        env_min = [0.0] + [min(d[str(r)][n][0] for r in env_runs)
                           for n, _ in hs]
        env_max = [0.0] + [max(d[str(r)][n][1] for r in env_runs)
                           for n, _ in hs]
        if is_exp:  # top restraint pin (wall crest / top beam)
            zs, env_min, env_max = zs + [2.85], env_min + [0.0], env_max + [0.0]
        axL.plot(env_min, zs, "k.-", ms=9, mfc="gold", mec="k", lw=1.4)
        axL.plot(env_max, zs, "k.-", ms=9, mfc="gold", mec="k", lw=1.4)
        axL.plot([m for m in env_min], zs, "o", ms=6, mfc="gold", mec="k")
        axL.plot([m for m in env_max], zs, "o", ms=6, mfc="gold", mec="k")
        axL.axvline(0, color="red", lw=0.8, alpha=0.6)
        if note:
            axL.text(0.03, 0.95, note, transform=axL.transAxes,
                     fontsize=8, va="top", color="0.35")
        axL.set_xlim(-35, 35); axL.set_ylim(0, 3)
        axL.set_title(ttl, fontsize=12, fontweight="bold")
        axL.set_ylabel("Height (m)", fontsize=11)

        # ---- right: normalized first vs last
        for rn, colr, lbl in ((first, "mediumblue", "First run"),
                              (last, "red", "Last run")):
            prof = [max(abs(v) for v in d[str(rn)][n]) for n, _ in hs]
            ref = prof[[n for n, _ in hs].index("top")]
            ref = ref if ref > 0 else 1.0
            xs = [0.0] + [p / ref for p in prof]
            zz = [0.0] + [z for _, z in hs]
            if is_exp:
                xs, zz = xs + [0.0], zz + [2.85]
            axR.plot(xs, zz, "o-", color=colr, ms=6, lw=1.4, label=lbl)
        axR.set_xlim(-0.05, 1.5); axR.set_ylim(0, 3)
        axR.set_title(ttl, fontsize=12, fontweight="bold")
        axR.legend(fontsize=9, loc="center right")
    axes[3, 0].set_xlabel("Max. horizontal displacement (mm)", fontsize=12)
    axes[3, 1].set_xlabel("Normalized horizontal displacement", fontsize=12)
    fig.tight_layout()
    fig.savefig(str(PP / "figP8_disp_profiles.png"), dpi=300,
                bbox_inches="tight")
    plt.close(fig); print("-> figP8_disp_profiles.png")


def hysteresis_fig(plt):
    import openpyxl
    fig, axes = plt.subplots(2, 2, figsize=(12.5, 9.5))
    for ax, rn in zip(axes.ravel(), HYST_RUNS):
        # exp US-1 (processed_globalzero): U_avg (m) vs base_shear_kN
        fe = PROC9 / ("Test9Run%d_processed_globalzero.xlsx" % rn)
        if fe.exists():
            wb = openpyxl.load_workbook(str(fe), read_only=True,
                                        data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            next(rows)
            u, bs = [], []
            for r in rows:
                if r[0] is None:
                    continue
                u.append(r[1]); bs.append(r[3])
            wb.close()
            u = np.asarray(u, float) * 1000.0
            bs = np.asarray(bs, float)
            ax.plot(u, bs, "-", color="royalblue", lw=0.7, alpha=0.85,
                    label="US-1 (U_avg, base shear)")
        # sim: U_top = mean(Ch3, Ch4) rel table (same definition as the
        # experimental U_avg), re-zeroed per run, vs base-joint shear sum
        folder = glob.glob(str(SIM / ("Run%02d_*" % rn)))
        if folder:
            f3 = glob.glob(folder[0] + "/*Channel_3_DispTopQLeft*")
            f4 = glob.glob(folder[0] + "/*Channel_4_DispTopQRight*")
            f5 = glob.glob(folder[0] + "/*Channel_5_DispTable*")
            fbs = glob.glob(folder[0] + "/*cstav*")
            if f3 and f4 and f5 and fbs:
                _, x3 = read_hist(f3[0])
                _, x4 = read_hist(f4[0])
                _, x5 = read_hist(f5[0])
                _, v = read_hist(fbs[0])
                n = min(len(x3), len(x4), len(x5), len(v))
                u = 0.5 * (x3[:n] + x4[:n]) - x5[:n]
                ax.plot((u - u[0]) * 1000.0, v[:n], "-", color="red",
                        lw=0.8, label="NUM (U_top, base-joint shear)")
        ax.axhline(0, color="0.4", lw=0.6)
        ax.axvline(0, color="0.4", lw=0.6)
        ax.set_title("Run %d" % rn, fontsize=12, fontweight="bold")
        ax.legend(fontsize=9, loc="upper right")
    for ax in axes[1]:
        ax.set_xlabel("Top OOP displacement (mm)", fontsize=12)
    for ax in axes[:, 0]:
        ax.set_ylabel("Base shear (kN)", fontsize=12)
    fig.tight_layout()
    fig.savefig(str(PP / "figP9_hysteresis.png"), dpi=300,
                bbox_inches="tight")
    plt.close(fig); print("-> figP9_hysteresis.png")


if __name__ == "__main__":
    st = load_state()
    mode = sys.argv[1]
    if mode == "exp9":
        do_exp(st, "exp9", sys.argv[2], "Test9Run",
               int(sys.argv[3]), int(sys.argv[4]))
    elif mode == "exp12":
        do_exp(st, "exp12", sys.argv[2], "Test12Run",
               int(sys.argv[3]), int(sys.argv[4]))
    elif mode == "sim":
        do_sim(st, SIM, "simprof")
    elif mode == "simratch":
        lo = int(sys.argv[2]) if len(sys.argv) > 2 else 1
        hi = int(sys.argv[3]) if len(sys.argv) > 3 else 25
        do_sim(st, SIM_RATCH, "simratchprof", lo, hi)
    elif mode == "fig":
        plt = style()
        profile_fig(plt, st)
        hysteresis_fig(plt)
    elif mode == "figoverlay":
        plt = style()
        overlay_fig(plt, st)
    elif mode == "figch3":
        plt = style()
        ch3_hysteresis_fig(plt)
