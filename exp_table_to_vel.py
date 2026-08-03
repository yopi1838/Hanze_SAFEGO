# -*- coding: ascii -*-
"""
Build a 3DEC velocity table from a measured shake-table record.
=============================================================================
Produces a vel_run_NN.txt in exactly the format strategy_C_3dec_*.py writes,
so it can be imported by the existing driver machinery without changes:

    line 1        table name
    line 2        N <tab> 0
    lines 3..N+2  time <tab> velocity-z  (s, m/s)

WHY THIS EXISTS
    Strategy C synthesises a 1.5-cycle sinusoid tuned to the wall's identified
    period, which stays near 0.12 s. That delivers acceleration but very little
    displacement: in stratC_results_GI_NORATCH the table moves 1.3-4.2 mm in
    runs 17-25, while the specimen's table moved 12-61 mm over the same runs
    (0.04-0.25x). A cracked URM wall is a rocking mechanism governed by
    displacement relative to its 210 mm thickness, so the synthetic pulse
    cannot drive it to collapse no matter how the joint model is tuned.
    This script feeds the model the motion the specimen actually saw.

SOURCE CHANNEL
    Channel 5, 'Disp - Shake table' (metres), column r[6] of the xlsx.
    Velocity is obtained by central-difference differentiation.

    Channel 12, 'Acc - Shake table' (g), is the obvious alternative but has to
    be integrated, and any accelerometer offset integrates into a linear
    velocity ramp and a quadratic displacement drift. Differentiating a
    displacement record has no such failure mode -- it only amplifies
    high-frequency noise, which --lowpass handles. Use --from-acc if you want
    the acceleration route for comparison.

CONDITIONING
    1. window to the strong-motion band (Arias 0.5-99.5% by default, or an
       explicit --t0/--t1) so the solve is seconds rather than ~100 s
    2. remove the mean velocity over the window, so the table has no net drift
    3. cosine taper both ends, so the model is not hit with a step in velocity
       from rest
    Each step's cost is reported: residual net table drift, and how much peak
    displacement the windowing discarded.

USAGE
    python exp_table_to_vel.py 24                     # US-1 (Test 9) run 24
    python exp_table_to_vel.py 24 --test 12           # US-2 (Test 12)
    python exp_table_to_vel.py 24 --out vel_exp_24.txt
    python exp_table_to_vel.py 24 --t0 15.0 --t1 33.0 --taper 0.3
    python exp_table_to_vel.py 24 --lowpass 25        # Hz, if dv/dt is noisy

Style note: Python-2/3 compatible (.format(), no f-strings) to match the rest
of the codebase.
"""
import sys
import argparse
import numpy as np

try:
    import safego_paths as sp
except ImportError:
    sp = None

TABLE_DISP_COL = 6      # channel 5, 'Disp - Shake table', metres
TABLE_ACC_COL = 13      # channel 12, 'Acc - Shake table', g
TIME_COL = 1
G = 9.80665


def read_run(path):
    """-> (t, table_disp_m, table_acc_g). Blank/short rows are skipped."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows)                                   # header
    t, d, a = [], [], []
    for r in rows:
        if len(r) <= TABLE_ACC_COL or r[TIME_COL] is None:
            continue
        t.append(r[TIME_COL])
        d.append(r[TABLE_DISP_COL])
        a.append(r[TABLE_ACC_COL])
    wb.close()
    return (np.asarray(t, dtype=float),
            np.asarray(d, dtype=float),
            np.asarray(a, dtype=float))


def arias_window(a, lo=0.005, hi=0.995):
    """Index bounds of the significant-duration band of an acceleration trace."""
    a = a - np.mean(a[:200])
    e = np.cumsum(a ** 2)
    if e[-1] <= 0:
        return 0, len(a) - 1
    e = e / e[-1]
    return int(np.searchsorted(e, lo)), int(np.searchsorted(e, hi))


def lowpass(x, dt, fc):
    """Zero-phase Butterworth, falling back to a moving average if SciPy is
    unavailable (the 3DEC-embedded interpreter does not ship SciPy)."""
    try:
        from scipy.signal import butter, filtfilt
        b, a = butter(4, fc / (0.5 / dt), btype="low")
        return filtfilt(b, a, x)
    except ImportError:
        n = max(1, int(round(1.0 / (fc * dt))))
        k = np.ones(n) / float(n)
        return np.convolve(x, k, mode="same")


def cosine_taper(n, dt, sec):
    """Ramp of `sec` seconds at each end, 1.0 in between."""
    w = np.ones(n)
    m = int(round(sec / dt))
    if m < 2 or 2 * m >= n:
        return w
    ramp = 0.5 * (1.0 - np.cos(np.pi * np.arange(m) / float(m)))
    w[:m] = ramp
    w[-m:] = ramp[::-1]
    return w


def build(t, d, a, args):
    dt = float(np.median(np.diff(t)))

    if args.t0 is not None or args.t1 is not None:
        i0 = 0 if args.t0 is None else int(np.searchsorted(t, args.t0))
        i1 = len(t) - 1 if args.t1 is None else int(np.searchsorted(t, args.t1))
        how = "explicit --t0/--t1"
    else:
        i0, i1 = arias_window(a)
        how = "Arias 0.5-99.5%"

    pp = lambda x: float(np.max(x) - np.min(x))
    d_full_pp = pp(d)
    d = d[i0:i1 + 1]
    t = t[i0:i1 + 1]
    a = a[i0:i1 + 1]

    if args.from_acc:
        acc = (a - np.mean(a[:max(10, int(0.1 / dt))])) * G
        v = np.concatenate(([0.0], np.cumsum(0.5 * (acc[1:] + acc[:-1]) * dt)))
        source = "channel 12 acceleration, integrated"
    else:
        v = np.gradient(d, dt)
        source = "channel 5 displacement, differentiated"

    if args.lowpass:
        v = lowpass(v, dt, args.lowpass)

    v = v - np.mean(v)                                   # no net drift
    v = v * cosine_taper(len(v), dt, args.taper)         # start/end at rest

    tt = t - t[0]
    d_recon = np.concatenate(([0.0], np.cumsum(0.5 * (v[1:] + v[:-1]) * dt)))

    # peak-to-peak is baseline-free, so it is the honest way to check that the
    # conditioned table still carries the motion the specimen actually saw
    info = {
        "dt": dt, "n": len(v), "dur": float(tt[-1]), "how": how,
        "source": source,
        "v_peak": float(np.max(np.abs(v))),
        "d_full_pp_mm": d_full_pp * 1000.0,
        "d_win_pp_mm": pp(d) * 1000.0,
        "d_recon_pp_mm": pp(d_recon) * 1000.0,
        "fidelity": pp(d_recon) / max(pp(d), 1e-12),
        "ramp_removed_mm": float(d[-1] - d[0]) * 1000.0,
        "d_drift_mm": float(d_recon[-1]) * 1000.0,
    }
    return tt, v, info


def write_table(path, name, t, v):
    with open(str(path), "w", newline="\n") as f:
        f.write("{}\n".format(name))
        f.write("{}\t0\n".format(len(t)))
        for ti, vi in zip(t, v):
            f.write("{:.6f}\t{:.9e}\n".format(ti, vi))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("run", type=int, help="experimental run number")
    ap.add_argument("--test", type=int, default=9, choices=(9, 12),
                    help="9 = US-1, 12 = US-2 (default 9)")
    ap.add_argument("--exp-dir", default=None,
                    help="EXP_DATA folder (default: safego_paths.exp_data_dir())")
    ap.add_argument("--out", default=None, help="output table file")
    ap.add_argument("--name", default=None, help="3DEC table name")
    ap.add_argument("--t0", type=float, default=None)
    ap.add_argument("--t1", type=float, default=None)
    ap.add_argument("--taper", type=float, default=0.25,
                    help="cosine taper at each end, seconds (default 0.25)")
    ap.add_argument("--lowpass", type=float, default=None,
                    help="low-pass corner in Hz applied to the velocity")
    ap.add_argument("--from-acc", action="store_true",
                    help="integrate channel 12 instead of differentiating "
                         "channel 5 (drift-prone, for comparison only)")
    args = ap.parse_args()

    from pathlib import Path
    if args.exp_dir:
        exp_dir = Path(args.exp_dir)
    elif sp is not None:
        exp_dir = sp.exp_data_dir()
    else:
        exp_dir = Path("EXP_DATA")

    src = exp_dir / "Test{}Run{}.xlsx".format(args.test, args.run)
    if not src.is_file():
        raise SystemExit("not found: {}".format(src))

    print("reading {}".format(src))
    t, d, a = read_run(src)
    tt, v, info = build(t, d, a, args)

    name = args.name or "exp_T{}_run{:02d}".format(args.test, args.run)
    out = Path(args.out) if args.out else Path(
        "vel_exp_T{}_run{:02d}.txt".format(args.test, args.run))
    write_table(out, name, tt, v)

    print("  source        : {}".format(info["source"]))
    print("  window        : {} -> {:.3f} s, {} pts at dt = {:.5f} s"
          .format(info["how"], info["dur"], info["n"], info["dt"]))
    print("  table motion  : {:.2f} mm peak-to-peak in the window "
          "({:.2f} mm over the full record)"
          .format(info["d_win_pp_mm"], info["d_full_pp_mm"]))
    print("  re-integrated : {:.2f} mm peak-to-peak  ->  fidelity {:.3f} "
          "(want ~1.000)".format(info["d_recon_pp_mm"], info["fidelity"]))
    print("  ramp removed  : {:+.2f} mm of instrument drift over the window"
          .format(info["ramp_removed_mm"]))
    print("  residual drift: {:+.3f} mm at the end (want ~0)"
          .format(info["d_drift_mm"]))
    print("  peak velocity : {:.4f} m/s".format(info["v_peak"]))
    print("  -> {}   (3DEC table name '{}')".format(out, name))
    print("")
    print("  In 3DEC, with this folder as cwd:")
    print("    table '{}' import '{}'".format(name, out.as_posix()))
    print("    block apply velocity-z 1.0 table '{}' range group 'S'".format(name))
    print("    block apply velocity-z 1.0 table '{}' range group 'T_B'".format(name))
    print("    model solve dynamic time {:.6f}".format(info["dur"]))
    print("    block gridpoint apply-remove velocity-z range group 'S'")
    print("    block gridpoint apply-remove velocity-z range group 'T_B'")
    print("    model solve dynamic time 2.5")


if __name__ == "__main__":
    main()
