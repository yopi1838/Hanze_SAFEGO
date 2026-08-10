# -*- coding: ascii -*-
"""
Build the Strategy D protocol table from the measured shake-table records.
=============================================================================
Strategy D replaces the spectrum-based intensity measure with two quantities
taken directly from each measured table motion:

    PGD    peak ground displacement, as an AMPLITUDE, half the peak-to-peak
           table displacement inside the significant-duration window   [m]
    PGV    peak ground velocity over the same window                   [m/s]

from which a single-sinusoid idealisation follows:

    T_eq = pi * PGD / PGV          equivalent period of the pulse      [s]
    A    = PGV * 2*pi / T_eq       implied peak acceleration           [m/s2]

WHY THESE TWO
    The intensity measure has to be a property of the ground motion alone,
    otherwise it is not comparable between runs, between records, or against
    any other study. Sd(record, T_current) failed that test because T_current
    is a property of the damaged structure, which produced the feedback loop
    documented in strategy_C_3dec_primed.py. PGD and PGV are record
    properties, so invariance holds by construction.

    Neither alone is sufficient. In US-1, run 3 has a LARGER table
    displacement than run 21 and did essentially nothing to the wall
    (0.02 mm against 16.8 mm), because its 1.11 s equivalent period puts it
    at 0.04 g. The pair (PGD, T_eq) captures both, and PGD is the quantity
    that governs rocking stability relative to the 210 mm wall thickness.

VALIDATION BUILT IN
    The implied PGA from the two-parameter idealisation is compared against
    the PGA measured on channel 12. Agreement to roughly 20% on the damaging
    runs is what justifies replacing a ~100 s record with a ~0.3 s pulse.
    Long-period, low-amplitude runs agree less well, because their measured
    PGA is dominated by high-frequency content the idealisation discards.
    Both columns are written out so the comparison is visible, not asserted.

OUTPUT
    protocol_D_US<N>.csv with one row per run:
        run, PGD_amp_mm, PGV_mps, T_eq_s, PGA_implied_g, PGA_measured_g,
        dur_window_s, EDP_peak_rel_mm
    EDP_peak_rel_mm is the measured wall response (mean of channels 3 and 4,
    already table-relative), carried through for convenience when plotting.

USAGE
    python make_protocol_D.py                 # US-1 (Test 9)
    python make_protocol_D.py --test 12       # US-2 (Test 12)
    python make_protocol_D.py --runs 1 24

Style note: Python-2/3 compatible (.format(), no f-strings).
"""
import argparse
import csv
import math
from pathlib import Path

import numpy as np

try:
    import safego_paths as sp
except ImportError:
    sp = None

TIME_COL = 1
CH3_COL = 4          # 'Disp - Top quarter left'   [m]
CH4_COL = 5          # 'Disp - Top quarter right'  [m]
TABLE_DISP_COL = 6   # 'Disp - Shake table'        [m]
TABLE_ACC_COL = 13   # 'Acc - Shake table'         [g]


def read_run(path):
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    t, d, a, c3, c4 = [], [], [], [], []
    for r in rows:
        if len(r) <= TABLE_ACC_COL or r[TIME_COL] is None:
            continue
        t.append(r[TIME_COL])
        d.append(r[TABLE_DISP_COL])
        a.append(r[TABLE_ACC_COL])
        c3.append(r[CH3_COL])
        c4.append(r[CH4_COL])
    wb.close()
    f = lambda x: np.asarray(x, dtype=float)
    return f(t), f(d), f(a), f(c3), f(c4)


def significant_window(x, lo=0.005, hi=0.995):
    """Significant-duration bounds from the energy of `x`.

    Pass the DISPLACEMENT trace, not the acceleration. An Arias-style window
    built on acceleration is dominated by high-frequency content and can miss
    a large low-frequency table excursion entirely. US-1 run 18 is the case
    that exposed this: its biggest table swing (17.97 mm at t = 11 s) fell
    outside the acceleration window, which reported 4.25 mm peak-to-peak and
    a PGD of 2.13 mm, sitting absurdly between run 17 at 8.08 and run 19 at
    10.03. On the displacement window it returns 8.99 mm, and every other run
    in the set is unchanged to the last decimal."""
    x = x - np.mean(x)
    e = np.cumsum(x ** 2)
    if e[-1] <= 0:
        return 0, len(x) - 1
    e = e / e[-1]
    return int(np.searchsorted(e, lo)), int(np.searchsorted(e, hi))


def analyse(t, d, a, c3, c4):
    dt = float(np.median(np.diff(t)))
    i0, i1 = significant_window(d)
    dw = d[i0:i1 + 1]
    vw = np.gradient(d, dt)[i0:i1 + 1]

    # PGD as an amplitude: T_eq = pi*PGD/PGV assumes a sinusoid of amplitude
    # PGD, so half the peak-to-peak is the consistent measure. Note this is
    # NOT the same convention as peak_table_mm in exp_Test9_metrics.csv, which
    # is peak relative to the record start over the whole trace.
    PGD = float(np.max(dw) - np.min(dw)) / 2.0
    PGV = float(np.max(np.abs(vw - np.mean(vw))))
    if PGD <= 0 or PGV <= 0:
        return None
    T_eq = math.pi * PGD / PGV
    A_impl = PGV * (2.0 * math.pi / T_eq)
    # PGA is a peak, so it needs no window; taking it over the whole record
    # avoids penalising the idealisation for content the window excluded.
    A_meas = float(np.max(np.abs(a - np.mean(a[:200])))) * 9.80665

    # measured wall response, mean of the two top-quarter sensors. They are
    # frame-mounted so already table-relative (CLAUDE.md section 9).
    top = 0.5 * (c3 + c4)
    edp = float(np.max(np.abs(top - top[0]))) * 1000.0

    return {
        "PGD_amp_mm": round(PGD * 1000.0, 4),
        "PGV_mps": round(PGV, 5),
        "T_eq_s": round(T_eq, 5),
        "PGA_implied_g": round(A_impl / 9.80665, 4),
        "PGA_measured_g": round(A_meas / 9.80665, 4),
        "dur_window_s": round(float(t[i1] - t[i0]), 3),
        "EDP_peak_rel_mm": round(edp, 4),
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--test", type=int, default=9, choices=(9, 12))
    ap.add_argument("--exp-dir", default=None)
    ap.add_argument("--runs", type=int, nargs=2, default=None,
                    metavar=("LO", "HI"))
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    if args.exp_dir:
        exp_dir = Path(args.exp_dir)
    elif sp is not None:
        exp_dir = sp.exp_data_dir()
    else:
        exp_dir = Path("EXP_DATA")

    lo, hi = args.runs if args.runs else (1, 25)
    out = Path(args.out) if args.out else Path(
        "protocol_D_US{}.csv".format(1 if args.test == 9 else 2))

    cols = ["run", "PGD_amp_mm", "PGV_mps", "T_eq_s", "PGA_implied_g",
            "PGA_measured_g", "dur_window_s", "EDP_peak_rel_mm"]
    rows = []
    for rn in range(lo, hi + 1):
        f = exp_dir / "Test{}Run{}.xlsx".format(args.test, rn)
        if not f.is_file():
            print("  missing, skipped: {}".format(f.name))
            continue
        res = analyse(*read_run(f))
        if res is None:
            print("  run {:2d}: unusable record, skipped".format(rn))
            continue
        res["run"] = rn
        rows.append(res)
        print("  run {:2d}: PGD {:7.2f} mm  PGV {:6.3f} m/s  T_eq {:6.3f} s  "
              "PGA {:5.2f} g (meas {:5.2f})  EDP {:7.2f} mm".format(
                  rn, res["PGD_amp_mm"], res["PGV_mps"], res["T_eq_s"],
                  res["PGA_implied_g"], res["PGA_measured_g"],
                  res["EDP_peak_rel_mm"]))

    with open(str(out), "w", newline="\n") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(cols)
        for r in rows:
            w.writerow([r[c] for c in cols])
    print("\n-> {}  ({} runs)".format(out, len(rows)))

    if rows:
        num = [r for r in rows if r["PGA_measured_g"] > 0]
        err = [100.0 * (r["PGA_implied_g"] - r["PGA_measured_g"])
               / r["PGA_measured_g"] for r in num]
        big = [r for r in num if r["PGA_measured_g"] >= 0.2]
        errb = [100.0 * (r["PGA_implied_g"] - r["PGA_measured_g"])
                / r["PGA_measured_g"] for r in big]
        print("idealisation error in PGA, all runs      : "
              "mean {:+.0f}%, worst {:+.0f}%".format(
                  float(np.mean(err)), max(err, key=abs)))
        if errb:
            print("                        runs above 0.2 g : "
                  "mean {:+.0f}%, worst {:+.0f}%".format(
                      float(np.mean(errb)), max(errb, key=abs)))


if __name__ == "__main__":
    main()
