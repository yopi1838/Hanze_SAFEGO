# -*- coding: ascii -*-
"""
Channel-19 cross-validation: abs max OOP displacement per run at the
top-quarter-right sensor position (y=+0.54 m, z=2.06 m).

  Simulation : Channel_19_DispTopQRight minus Channel_5_DispTable
               (block histories are absolute; the experiment sensor is
               table-mounted, so the sim must be referenced to the table)
  Experiment : channel 4 'Disp - Top quarter right' from TestNRunM.xlsx
               (already wall-relative; baselined to the run start)

  peak = max |x(t) - x(t0)| within the run, in mm.

Outputs (into <SIM>/postproc/):
  xval_ch19_peaks.csv        run, sim, US-1, US-2 peaks (mm)
  figP6_ch19_peak_disp.png   paper-style comparison figure

Usage:
  python ch19_xval.py sim                     # simulation extraction
  python ch19_xval.py exp9  EXP_DIR LO HI     # US-1 runs LO..HI (staged)
  python ch19_xval.py exp12 EXP_DIR LO HI     # US-2 runs LO..HI (staged)
  python ch19_xval.py finish                  # write CSV + figure
EXP_DIR is normally ./EXP_DATA (override: SAFEGO_EXP_DATA).
State is cached in ./.cache/ch19_state.json so stages can run separately.
"""
import sys, json, csv, glob
import numpy as np
from pathlib import Path

import safego_paths as sp

HERE = sp.ROOT
SIM = sp.sim_dir()
SIM_DAMP = sp.sim_dir("stratC_results_GI_NORATCH")
SIM_RATCH = sp.sim_dir("stratC_results_RATCHETING")
PP = sp.postproc_dir()
STATE = sp.state_file("ch19")

def load_state():
    if STATE.exists():
        return json.loads(STATE.read_text())
    return {"sim": {}, "simdamp": {}, "exp9": {}, "exp12": {}}

def save_state(st):
    STATE.write_text(json.dumps(st))

def read_hist(fpath):
    d = np.genfromtxt(str(fpath), skip_header=2)
    m = np.isfinite(d[:, 0]) & np.isfinite(d[:, 1])
    return d[m, 0], d[m, 1]

def do_sim(st, sim_dir=None, key="sim", subtract_table=True):
    st.setdefault(key, {})
    sim_dir = SIM if sim_dir is None else sim_dir
    for folder in sorted(sim_dir.iterdir()):
        if not (folder.is_dir() and folder.name.startswith("Run")):
            continue
        rn = int(folder.name[3:5])
        f19 = glob.glob(str(folder / "*Channel_19*"))
        f05 = glob.glob(str(folder / "*Channel_5_DispTable*"))
        if not (f19 and f05):
            continue
        t19, c19 = read_hist(f19[0])
        if subtract_table:
            t05, c05 = read_hist(f05[0])
            n = min(len(c19), len(c05))
            rel = c19[:n] - c05[:n]
        else:
            rel = c19
        st[key][str(rn)] = round(float(np.max(np.abs(rel - rel[0]))) * 1000, 3)
        print("%s run %2d: %.2f mm" % (key, rn, st[key][str(rn)]))

def do_exp(st, key, exp_dir, prefix, lo, hi):
    import openpyxl
    for rn in range(lo, hi + 1):
        f = Path(exp_dir) / ("%s%d.xlsx" % (prefix, rn))
        if not f.exists():
            print("missing:", f.name); continue
        wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        next(rows)
        v = [r[5] for r in rows if r[1] is not None]  # col 5 = channel 4
        wb.close()
        x = np.asarray(v, dtype=float)
        st[key][str(rn)] = round(float(np.max(np.abs(x - x[0]))) * 1000, 3)
        print("%s run %2d: %.2f mm" % (key, rn, st[key][str(rn)]))
        save_state(st)

def do_finish(st):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    plt.rcParams.update({
        "font.family": "serif", "font.serif": ["DejaVu Serif"],
        "mathtext.fontset": "dejavuserif", "axes.linewidth": 1.2,
        "xtick.direction": "in", "ytick.direction": "in",
        "xtick.top": True, "ytick.right": True,
        "xtick.minor.visible": True, "ytick.minor.visible": True,
        "axes.grid": True, "grid.linestyle": ":", "grid.alpha": 0.6,
        "legend.frameon": True, "legend.framealpha": 1.0,
        "legend.edgecolor": "0.7"})

    runs = sorted(set(int(k) for d in st.values() for k in d))
    with open(str(PP / "xval_ch19_peaks.csv"), "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["run", "sim_nodamp_rel_mm", "sim_nodamp_abs_mm",
                    "sim_rayleigh1p5_rel_mm", "sim_ratch3pct_abs_mm",
                    "US1_peak_mm", "US2_peak_mm"])
        for rn in runs:
            w.writerow([rn, st["sim"].get(str(rn), ""),
                        st.get("sim_abs", {}).get(str(rn), ""),
                        st.get("simdamp", {}).get(str(rn), ""),
                        st.get("simratch_abs", {}).get(str(rn), ""),
                        st["exp9"].get(str(rn), ""),
                        st["exp12"].get(str(rn), "")])
    print("-> xval_ch19_peaks.csv")

    def series(d):
        rn = sorted(int(k) for k in d)
        return rn, [d[str(r)] for r in rn]

    fig, ax = plt.subplots(figsize=(9.2, 5.0))
    r1, y1 = series(st["exp9"])
    ax.plot(r1, y1, "o--", color="royalblue", ms=6, lw=1.2,
            label="US-1 (Test 9, Ch 4 top quarter right)")
    r2, y2 = series(st["exp12"])
    ax.plot(r2, y2, "o--", color="orange", ms=6, lw=1.2,
            label="US-2 (Test 12, Ch 4 top quarter right)")
    rs, ys = series(st.get("sim_abs") or st["sim"])
    rs = np.array(rs); ys = np.array(ys)
    YCAP = 35.0
    ax.plot(rs, np.minimum(ys, YCAP), "s-", color="red", ms=6, lw=1.6,
            label="NUM (Ch 19, not table-referenced)"
                  if st.get("sim_abs") else "NUM (Ch 19 rel. to table)",
            zorder=4)
    off = [(r, v) for r, v in zip(rs, ys) if v > YCAP]
    for k, (rr, v) in enumerate(off):
        ax.plot(rr, YCAP, "s", ms=8, mfc="white", mec="red", mew=1.6, zorder=5)
        ax.annotate("Run %d: %.0f mm" % (rr, v), xy=(rr, YCAP),
                    xytext=(rr - 9.5, YCAP - 3.5 - 3.2 * k), fontsize=10,
                    color="red",
                    arrowprops=dict(arrowstyle="->", color="red", lw=1.0))
    ax.set_xlim(0, 26); ax.set_ylim(-2, YCAP + 1)
    ax.set_xticks(range(0, 27, 2))
    ax.set_xlabel("Run number", fontsize=14)
    ax.set_ylabel("Abs max OOP displacement (mm)", fontsize=14)
    ax.legend(fontsize=11, loc="upper left")
    fig.tight_layout()
    fig.savefig(str(PP / "figP6_ch19_peak_disp.png"), dpi=300,
                bbox_inches="tight")
    plt.close(fig)
    print("-> figP6_ch19_peak_disp.png")

    # figP7: damping comparison (no viscous damping vs ratcheting run)
    if st.get("simratch_abs"):
        fig, ax = plt.subplots(figsize=(9.2, 5.0))
        r1, y1 = series(st["exp9"])
        ax.plot(r1, y1, "o--", color="royalblue", ms=6, lw=1.2,
                label="US-1 (Test 9)")
        r2, y2 = series(st["exp12"])
        ax.plot(r2, y2, "o--", color="orange", ms=6, lw=1.2,
                label="US-2 (Test 12)")
        for key, colr, mk, lbl in (
                ("sim_abs", "red", "s",
                 "NUM: no viscous damping (hysteretic only)"),
                ("simratch_abs", "purple", "d",
                 "NUM: 3% Rayleigh, asym. pulse (v5 model)")):
            rs, ys = series(st[key])
            rs = np.array(rs); ys = np.array(ys)
            ax.plot(rs, np.minimum(ys, YCAP), mk + "-", color=colr, ms=6,
                    lw=1.6, label=lbl, zorder=4)
            off = [(r, v) for r, v in zip(rs, ys) if v > YCAP]
            for k, (rr, v) in enumerate(off):
                ax.plot(rr, YCAP, mk, ms=8, mfc="white", mec=colr, mew=1.6,
                        zorder=5)
                ax.annotate("Run %d: %.0f mm" % (rr, v), xy=(rr, YCAP),
                            xytext=(rr - 9.5, YCAP - 3.5 - 3.2 * k),
                            fontsize=10, color=colr,
                            arrowprops=dict(arrowstyle="->", color=colr,
                                            lw=1.0))
        ax.set_xlim(0, 26); ax.set_ylim(-2, YCAP + 1)
        ax.set_xticks(range(0, 27, 2))
        ax.set_xlabel("Run number", fontsize=14)
        ax.set_ylabel("Abs max OOP displacement (mm)", fontsize=14)
        ax.legend(fontsize=10, loc="upper left")
        fig.tight_layout()
        fig.savefig(str(PP / "figP7_ch19_damping_compare.png"), dpi=300,
                    bbox_inches="tight")
        plt.close(fig)
        print("-> figP7_ch19_damping_compare.png")

if __name__ == "__main__":
    st = load_state()
    mode = sys.argv[1] if len(sys.argv) > 1 else "sim"
    if mode == "sim":
        do_sim(st); save_state(st)
    elif mode == "simdamp":
        do_sim(st, SIM_DAMP, "simdamp"); save_state(st)
    elif mode == "simabs":
        do_sim(st, SIM, "sim_abs", subtract_table=False); save_state(st)
    elif mode == "simratchabs":
        do_sim(st, SIM_RATCH, "simratch_abs", subtract_table=False)
        save_state(st)
    elif mode == "exp9":
        do_exp(st, "exp9", sys.argv[2], "Test9Run",
               int(sys.argv[3]), int(sys.argv[4]))
    elif mode == "exp12":
        do_exp(st, "exp12", sys.argv[2], "Test12Run",
               int(sys.argv[3]), int(sys.argv[4]))
    elif mode == "finish":
        do_finish(st)
