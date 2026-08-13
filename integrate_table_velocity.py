# -*- coding: ascii -*-
"""
Shake-table acceleration -> velocity, for the HU / EC / FR input signals.
=============================================================================
Reads the shake-table acceleration channel (channel 12, "Acc - Shake table",
recorded in g) from each signal file, integrates it once to velocity, and plots
acceleration and velocity against time for the three records side by side. The
cumulative Arias intensity (Husid curve) is overlaid on each acceleration panel
on a secondary axis, with the 5-95% significant-duration band highlighted.

ARIAS INTENSITY
    I_A(t) = (pi / 2g) * integral_0^t a(tau)^2 d tau            [m/s]
    The total I_A is a single-number measure of the energy of the motion; the
    time to go from 5% to 95% of it is the Trifunac-Brady significant duration
    D_5-95, i.e. the length of the strong-motion phase.

COLUMN CONVENTION
    channel 12 = shake-table acceleration. Where it sits depends on the file:
      * CSV exported with the leading blank column dropped:  time = col 0,
        channels 1..N = cols 1..N, so channel 12 = index 12   <- default here
      * raw Test{T}Run{N}.xlsx:  blank = col 0, Time = col 1, channels start at
        col 2, so channel 12 = index 13
    Set TIME_COL / ACC_COL below to match your files (the .xlsx branch overrides
    them automatically).

INTEGRATION
    A raw accelerometer trace has a small DC offset that, once integrated,
    becomes a linear velocity ramp. Two standard corrections are applied
    (both toggle-able):
      1. DEMEAN     - subtract the mean acceleration (estimated on the quiet
                      lead-in) before integrating, killing the ramp at source.
      2. DETREND_VEL- remove any residual linear trend from the velocity, so it
                      starts and ends near zero.
    LOWPASS_HZ optionally low-passes the acceleration first if the record is
    noisy. Integration is cumulative-trapezoidal.

USAGE
    1. Point the SIGNALS list at your three 100%-intensity files.
    2. python integrate_table_accel.py

OUTPUTS
    table_accel_velocity.png     accel + velocity + Arias, truncation marked
    vel_<REC>.txt                truncated velocity as a 3DEC table, one per
                                 record (import + apply as a velocity BC)
    table_pgv_manifest.csv       record, table_file, PGV, PGA, Arias, D5-95,
                                 cut time, n_samples, dt, duration

Only needs numpy + matplotlib (+ openpyxl if you feed it .xlsx files).
"""
import os
import csv
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

# --------------------------------------------------------------------- CONFIG
# label, path to the 100%-intensity signal file (.csv or .xlsx)
SIGNALS = [
    ("HU", "INPUT/HU_100.csv"),
    ("EC", "INPUT/EC_100.csv"),
    ("FR", "INPUT/FR_100.csv"),
]

TIME_COL = 0        # column index of time      (CSV convention)
ACC_COL  = 12       # column index of channel 12 (shake-table acceleration)
ACC_UNIT = "g"      # "g" -> multiplied by 9.80665; use "m/s2" to skip

DEMEAN      = True   # remove the acceleration DC offset before integrating
DETREND_VEL = True   # remove a residual linear trend from the velocity
LOWPASS_HZ  = None   # e.g. 25.0 to low-pass the acceleration first, or None

# Truncation: keep the record from t = 0 up to the 95% Arias time, but SNAP the
# cut to the closest velocity zero-crossing so the kept motion ends at rest
# (no velocity step). This is the significant-duration window used for the
# 3DEC record tables.
TRUNCATE_TO_ARIAS = True
ARIAS_HI          = 0.95

# 3DEC velocity-table export. Writes the truncated velocity time-history of each
# record in 3DEC table format, plus a manifest listing PGV and the other
# intensity measures. In 3DEC:
#     table '<name>' import '<name>.txt'
#     block apply velocity-z 1.0 table '<name>' range group 'S'
WRITE_3DEC_TABLE = True
TABLE_DIR        = "velocity_output"  # output folder for the tables + manifest
TABLE_PREFIX     = "vel_"     # -> vel_HU.txt, vel_EC.txt, vel_FR.txt
CLAMP_ENDPOINTS  = True       # force v[0] = v[end] = 0 (start / end at rest)

CSV_DELIM   = ","    # CSV delimiter
CSV_SKIP    = 1      # header rows to skip in a CSV

# record colours follow the project convention (HU blue, EC orange, FR red)
COLOR = {"HU": "#1f77b4", "EC": "#ff7f0e", "FR": "#d62728"}
G = 9.80665


# ------------------------------------------------------------------- reading
def read_signal(path):
    """-> (t [s], a [as recorded]). Auto-detects .xlsx vs delimited text and
    uses the appropriate column convention."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        # raw xlsx: blank col 0, Time col 1, channel 12 at index 13
        t_col, a_col = 1, 13
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        next(rows)                                   # header
        t, a = [], []
        for r in rows:
            if len(r) <= a_col or r[t_col] is None:
                continue
            t.append(r[t_col]); a.append(r[a_col])
        wb.close()
        return np.asarray(t, float), np.asarray(a, float)
    # delimited text (CSV): use the configured columns
    data = np.genfromtxt(path, delimiter=CSV_DELIM, skip_header=CSV_SKIP)
    return data[:, TIME_COL].astype(float), data[:, ACC_COL].astype(float)


def lowpass(x, dt, fc):
    """Zero-phase Butterworth if SciPy is present, else a moving average."""
    try:
        from scipy.signal import butter, filtfilt
        b, a = butter(4, fc / (0.5 / dt), btype="low")
        return filtfilt(b, a, x)
    except ImportError:
        n = max(1, int(round(1.0 / (fc * dt))))
        return np.convolve(x, np.ones(n) / n, mode="same")


def arias_cumulative(t, a_ms2):
    """Cumulative Arias intensity I_A(t) = (pi/2g) * integral a^2 dt  [m/s]."""
    dt = float(np.median(np.diff(t)))
    sq = a_ms2 ** 2
    ia = np.concatenate(([0.0], np.cumsum(0.5 * (sq[1:] + sq[:-1]) * dt)))
    return (np.pi / (2.0 * G)) * ia


def significant_duration(t, ia, lo=0.05, hi=0.95):
    """(i5, i95, D_5-95) from a cumulative Arias curve."""
    tot = ia[-1]
    if tot <= 0:
        return 0, len(ia) - 1, 0.0
    i5 = int(np.searchsorted(ia, lo * tot))
    i95 = int(np.searchsorted(ia, hi * tot))
    i5 = min(i5, len(t) - 1); i95 = min(i95, len(t) - 1)
    return i5, i95, float(t[i95] - t[i5])


def write_3dec_table(path, name, t, v):
    """3DEC velocity table: line 1 = table name, line 2 = 'N <tab> 0', then
    'time <tab> velocity' rows. LF line endings, so it imports on Windows and
    Linux alike."""
    with open(path, "w", newline="\n") as f:
        f.write("{}\n".format(name))
        f.write("{}\t0\n".format(len(t)))
        for ti, vi in zip(t, v):
            f.write("{:.6f}\t{:.9e}\n".format(ti, vi))


def zero_vel_cut(v, i_target):
    """Index of the velocity zero-crossing closest to i_target. Snapping the
    truncation here means the kept record ends at zero table velocity -- at
    rest, with no velocity step."""
    s = np.sign(v)
    s[s == 0] = 1.0
    zc = np.where(np.diff(s) != 0)[0]        # v[i], v[i+1] straddle zero
    if len(zc) == 0:
        return len(v) - 1
    return int(zc[np.argmin(np.abs(zc - i_target))])


def to_velocity(t, a_raw):
    """Condition the acceleration and integrate it once to velocity.
    Returns (a_ms2, v_ms, info)."""
    dt = float(np.median(np.diff(t)))
    a = a_raw * (G if ACC_UNIT.lower() == "g" else 1.0)      # -> m/s2

    if LOWPASS_HZ:
        a = lowpass(a, dt, LOWPASS_HZ)
    if DEMEAN:
        lead = max(10, int(0.5 / dt))                        # ~first 0.5 s
        a = a - np.mean(a[:lead])

    # cumulative-trapezoidal integration a -> v
    v = np.concatenate(([0.0], np.cumsum(0.5 * (a[1:] + a[:-1]) * dt)))

    if DETREND_VEL and len(v) > 2:
        trend = np.polyval(np.polyfit(t, v, 1), t)           # linear baseline
        v = v - trend

    info = {
        "dt": dt, "dur": float(t[-1] - t[0]),
        "pga_g": float(np.max(np.abs(a))) / G,
        "pga_ms2": float(np.max(np.abs(a))),
        "pgv_ms": float(np.max(np.abs(v))),
        "drift_end": float(v[-1]),
    }
    return a, v, info


# ------------------------------------------------------------------- plotting
def main():
    results = []
    manifest = []
    if WRITE_3DEC_TABLE and not os.path.isdir(TABLE_DIR):
        os.makedirs(TABLE_DIR)
    for label, path in SIGNALS:
        if not os.path.isfile(path):
            print("  ! not found, skipped: {}".format(path))
            continue
        t, a_raw = read_signal(path)
        a, v, info = to_velocity(t, a_raw)
        ia = arias_cumulative(t, a)
        i5, i95, d595 = significant_duration(t, ia)
        info["arias_ms"] = float(ia[-1])
        info["D5_95"] = d595
        if TRUNCATE_TO_ARIAS:
            i_hi = int(np.searchsorted(ia, ARIAS_HI * ia[-1]))
            i_hi = min(max(i_hi, 1), len(t) - 1)
            icut = zero_vel_cut(v, i_hi)
        else:
            icut = len(t) - 1
        info["t95"] = float(t[i95])
        info["t_cut"] = float(t[icut])
        info["dur_cut"] = float(t[icut] - t[0])
        info["v_at_cut"] = float(v[icut])
        results.append((label, t, a, v, info, ia, i5, i95, icut))
        print("  {}: {:.1f} s | PGA={:.3f} g | PGV={:.3f} m/s | Ia={:.3f} m/s | "
              "D5-95={:.1f} s | cut @ {:.2f}s (t95={:.2f}s, v={:+.1e} m/s) -> "
              "kept {:.2f}s".format(
                  label, info["dur"], info["pga_g"], info["pgv_ms"],
                  info["arias_ms"], info["D5_95"], info["t_cut"], info["t95"],
                  info["v_at_cut"], info["dur_cut"]))

        # --- 3DEC velocity table (truncated velocity, starts at t=0) ---------
        if WRITE_3DEC_TABLE:
            tt = t[:icut + 1] - t[0]
            vv = v[:icut + 1].copy()
            if CLAMP_ENDPOINTS:
                vv[0] = 0.0
                vv[-1] = 0.0
            name = "{}{}".format(TABLE_PREFIX, label)
            tbl = os.path.join(TABLE_DIR, name + ".txt")
            write_3dec_table(tbl, name, tt, vv)
            info["table_file"] = os.path.basename(tbl)
            manifest.append({
                "record": label,
                "table_file": os.path.basename(tbl),
                "n_samples": len(tt),
                "dt": round(info["dt"], 6),
                "dur_s": round(float(tt[-1]), 4),
                "PGA_g": round(info["pga_g"], 5),
                "PGV_mps": round(info["pgv_ms"], 6),
                "Arias_mps": round(info["arias_ms"], 6),
                "D5_95_s": round(info["D5_95"], 4),
                "t_cut_s": round(info["t_cut"], 4),
            })
            print("       -> {}  (PGV = {:.4f} m/s, {} pts)".format(
                os.path.basename(tbl), info["pgv_ms"], len(tt)))
    if not results:
        raise SystemExit("no signal files found -- edit the SIGNALS list")

    if WRITE_3DEC_TABLE and manifest:
        man = os.path.join(TABLE_DIR, "table_pgv_manifest.csv")
        cols = ["record", "table_file", "n_samples", "dt", "dur_s", "PGA_g",
                "PGV_mps", "Arias_mps", "D5_95_s", "t_cut_s"]
        with open(man, "w", newline="\n") as f:
            w = csv.writer(f, lineterminator="\n")
            w.writerow(cols)
            for m in manifest:
                w.writerow([m[c] for c in cols])
        print("\n-> {}  ({} records)".format(man, len(manifest)))

    n = len(results)
    fig, axes = plt.subplots(n, 2, figsize=(13, 3.0 * n), squeeze=False)
    ARIAS_C = "#333333"
    CUT_C = "#7d1128"
    for i, (label, t, a, v, info, ia, i5, i95, icut) in enumerate(results):
        c = COLOR.get(label, "C{}".format(i))
        axL, axR = axes[i, 0], axes[i, 1]
        kept = slice(0, icut + 1)

        # acceleration on the primary axis (kept bold, discarded faint)
        if TRUNCATE_TO_ARIAS:
            axL.plot(t, a, color=c, lw=0.5, alpha=0.22, zorder=1)
            axL.plot(t[kept], a[kept], color=c, lw=0.7, zorder=2)
            axL.axvline(t[icut], color=CUT_C, lw=1.2, zorder=5)
        else:
            axL.plot(t, a, color=c, lw=0.7, zorder=2)
        axL.axhline(0, color="0.7", lw=0.6)
        axL.set_ylabel("{}\naccel. (m/s$^2$)".format(label), fontsize=10,
                       color=c)
        ttl = ("PGA = {:.3f} g    $I_A$ = {:.3f} m/s    "
               "$D_{{5\\text{{-}}95}}$ = {:.1f} s").format(
                   info["pga_g"], info["arias_ms"], info["D5_95"])
        if TRUNCATE_TO_ARIAS:
            ttl += "    | cut @ {:.1f} s".format(info["t_cut"])
        axL.set_title(ttl, fontsize=9)
        axL.grid(alpha=0.25)

        # cumulative Arias intensity (Husid curve) on a secondary axis
        axA = axL.twinx()
        axA.plot(t, ia, color=ARIAS_C, lw=1.3, ls="--", zorder=3)
        axA.plot(t[i5:i95 + 1], ia[i5:i95 + 1], color=ARIAS_C, lw=2.6,
                 zorder=4)                                  # 5-95% strong phase
        for k in (i5, i95):                                 # t5 / t95 markers
            axA.axvline(t[k], color=ARIAS_C, lw=0.7, ls=":", alpha=0.6)
        axA.set_ylabel("Arias $I_A$ (m/s)", fontsize=9, color=ARIAS_C)
        axA.tick_params(axis="y", labelcolor=ARIAS_C)
        axA.set_ylim(0, max(ia[-1] * 1.05, 1e-9))
        axA.margins(x=0)

        if TRUNCATE_TO_ARIAS:
            axR.plot(t, v, color=c, lw=0.5, alpha=0.22)
            axR.plot(t[kept], v[kept], color=c, lw=0.8)
            axR.axvline(t[icut], color=CUT_C, lw=1.2)
            axR.plot(t[icut], v[icut], "o", color=CUT_C, ms=5)  # zero-vel cut
        else:
            axR.plot(t, v, color=c, lw=0.8)
        axR.axhline(0, color="0.7", lw=0.6)
        axR.set_ylabel("velocity (m/s)", fontsize=10)
        ttlR = "PGV = {:.3f} m/s".format(info["pgv_ms"])
        if TRUNCATE_TO_ARIAS:
            ttlR += "    | kept {:.1f} s (ends at v=0)".format(info["dur_cut"])
        axR.set_title(ttlR, fontsize=9)
        axR.grid(alpha=0.25)

        if i == n - 1:
            axL.set_xlabel("Time (s)"); axR.set_xlabel("Time (s)")

    fig.suptitle("Shake-table acceleration and integrated velocity "
                 "(channel 12, 100% intensity)", fontsize=13, fontweight="bold")
    fig.tight_layout(rect=(0, 0, 1, 0.98))
    out = "table_accel_velocity.png"
    fig.savefig(out, dpi=200, bbox_inches="tight")
    plt.close(fig)
    print("\n-> {}".format(out))


if __name__ == "__main__":
    main()