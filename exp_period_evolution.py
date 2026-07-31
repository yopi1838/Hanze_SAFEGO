# -*- coding: utf-8 -*-
"""
Experimental period evolution per run via PSD analysis
=======================================================
Replicates the methodology from "Period Elongation Calculations Using PSD"
(Moshfeghi et al.) for all 25 runs of Test 9 (US1).
(Provided by Yopi; used via exp_period_wrapper.py in this environment.)
"""

import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from pathlib import Path
from scipy import signal
import csv, sys

HAS_PLOTLIB = False

EXP_DIR = Path(r"dynamic\Hanze\EXP_DATA_2")
OUT_DIR = Path(r"dynamic\results\Hanze")

COL_TIME = 1
COL_TABLE_DISP = 6
COL_ACC_TABLE = 13
COL_ACC_BOT = 16
COL_ACC_MID = 17
COL_ACC_TOP = 18

FILTER_LOWCUT = 0.25
FILTER_HIGHCUT = 50.0
FILTER_ORDER = 3
WELCH_WINDOW_SEC = 30.0
WELCH_OVERLAP_FRAC = 0.50
PSD_SMOOTHING_PTS = 9

MODE1_F_MIN = 5.0
MODE1_F_MAX = 15.0
MODE2_F_MIN = 16.0
MODE2_F_MAX = 22.0

PROTOCOL = [
    ( 1, "HU12", 0.50),  ( 2, "HU12", 0.75),  ( 3, "EC40", 0.20),
    ( 4, "HU12", 1.00),  ( 5, "HU12", 1.25),  ( 6, "EC40", 0.30),
    ( 7, "HU12", 1.50),  ( 8, "EC40", 0.40),  ( 9, "HU12", 1.75),
    (10, "HU12", 2.00),  (11, "EC40", 0.50),  (12, "HU12", 2.25),
    (13, "HU12", 2.50),  (14, "HU12", 2.75),  (15, "HU12", 3.00),
    (16, "HU12", 3.50),  (17, "HU12", 4.00),  (18, "HU12", 4.50),
    (19, "HU12", 5.00),  (20, "HU12", 5.50),  (21, "HU12", 6.00),
    (22, "FR76", 1.00),  (23, "FR76", 1.50),  (24, "FR76", 1.75),
    (25, "FR76", 2.00),
]


def load_exp_run(run_no, exp_dir, test_no=9):
    fname = Path(exp_dir) / "Test{}Run{}.xlsx".format(test_no, run_no)
    if not fname.exists():
        return None
    import openpyxl
    wb = openpyxl.load_workbook(str(fname), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        vals = [v if v is not None else 0.0 for v in row]
        data.append(vals)
    wb.close()
    data = np.array(data, dtype=float)
    ncols = data.shape[1]
    result = {
        't': data[:, COL_TIME],
        'table_disp': data[:, COL_TABLE_DISP] if ncols > COL_TABLE_DISP else None,
    }
    for key, col in [('acc_table', COL_ACC_TABLE), ('acc_bot', COL_ACC_BOT),
                     ('acc_mid', COL_ACC_MID), ('acc_top', COL_ACC_TOP)]:
        result[key] = data[:, col] if ncols > col else None
    return result


def find_free_vibration_start(t, table_disp, acc_table=None):
    dt = float(np.median(np.diff(t)))
    sig = acc_table if acc_table is not None else table_disp
    if sig is None:
        return t[0] + 0.6 * (t[-1] - t[0])
    sig_dm = sig - np.mean(sig)
    window_n = max(1, int(1.0 / dt))
    n_windows = len(sig_dm) // window_n
    rms_t, rms_v = [], []
    for i in range(n_windows):
        start = i * window_n
        end = start + window_n
        rms_t.append(t[start + window_n // 2])
        rms_v.append(np.sqrt(np.mean(sig_dm[start:end]**2)))
    rms_t = np.array(rms_t); rms_v = np.array(rms_v)
    if len(rms_v) == 0:
        return t[0] + 0.6 * (t[-1] - t[0])
    peak_rms = np.max(rms_v)
    threshold = 0.02 * peak_rms
    i_peak_rms = np.argmax(rms_v)
    below = rms_v < threshold
    first_stable = None
    for i in range(i_peak_rms, len(below) - 3):
        if all(below[i:i + 3]):
            first_stable = i
            break
    if first_stable is not None:
        t_fv = rms_t[first_stable]
    else:
        above = np.where(rms_v[i_peak_rms:] > threshold)[0]
        if len(above) > 0:
            t_fv = rms_t[i_peak_rms + above[-1]] + 3.0
        else:
            t_fv = rms_t[i_peak_rms] + 10.0
    t_fv = t_fv + 5.0
    return min(float(t_fv), float(t[-1] - 10.0))


def process_acceleration(t, acc, t_free_start):
    mask = t >= t_free_start
    t_fv = t[mask]
    acc_fv = acc[mask].copy()
    if len(acc_fv) < 100:
        return None, None, None
    dt = float(np.median(np.diff(t_fv)))
    fs = 1.0 / dt
    acc_fv = acc_fv - np.mean(acc_fv)
    nyq = 0.5 * fs
    low = FILTER_LOWCUT / nyq
    high = min(FILTER_HIGHCUT / nyq, 0.99)
    if low < high and low > 0:
        b, a = signal.butter(FILTER_ORDER, [low, high], btype='band')
        acc_fv = signal.filtfilt(b, a, acc_fv)
    return t_fv, acc_fv, dt


def compute_psd(acc, dt, window_sec=None):
    fs = 1.0 / dt
    if window_sec is None:
        window_sec = WELCH_WINDOW_SEC
    nperseg = int(window_sec * fs)
    nperseg = min(nperseg, len(acc))
    noverlap = int(nperseg * WELCH_OVERLAP_FRAC)
    freqs, psd = signal.welch(acc, fs=fs, nperseg=nperseg,
                              noverlap=noverlap, window='hann',
                              scaling='density')
    if PSD_SMOOTHING_PTS > 1:
        kernel = np.ones(PSD_SMOOTHING_PTS) / PSD_SMOOTHING_PTS
        psd = np.convolve(psd, kernel, mode='same')
    psd_db = 10 * np.log10(psd + 1e-30)
    return freqs, psd_db


def pick_peak(freqs, psd_db, f_min, f_max, min_prominence_db=3.0):
    mask = (freqs >= f_min) & (freqs <= f_max)
    if not np.any(mask):
        return None, None
    idx_band = np.where(mask)[0]
    psd_band = psd_db[idx_band]
    i_peak_in_band = np.argmax(psd_band)
    i_peak = idx_band[i_peak_in_band]
    f_peak = float(freqs[i_peak])
    psd_peak = float(psd_db[i_peak])
    median_level = float(np.median(psd_band))
    if psd_peak - median_level < min_prominence_db:
        return None, None
    if i_peak > 0 and i_peak < len(freqs) - 1:
        y0 = psd_db[i_peak - 1]; y1 = psd_db[i_peak]; y2 = psd_db[i_peak + 1]
        denom = y0 - 2*y1 + y2
        if abs(denom) > 1e-12:
            df = float(freqs[1] - freqs[0])
            f_peak = f_peak + 0.5 * (y0 - y2) / denom * df
    return f_peak, psd_peak


def analyze_run(run_no, exp_dir, test_no=9, save_psd_plots=False, out_dir=None):
    data = load_exp_run(run_no, exp_dir, test_no=test_no)
    if data is None:
        return None
    result = {'run': run_no}
    t_fv_start = find_free_vibration_start(
        data['t'], data.get('table_disp'), data.get('acc_table'))
    result['t_free_vib_start'] = t_fv_start
    result['record_duration'] = float(data['t'][-1] - data['t'][0])
    dt = float(np.median(np.diff(data['t'])))
    result['fs'] = 1.0 / dt

    channels = {'acc_table': 'Shake table', 'acc_bot': 'Bottom quarter',
                'acc_mid': 'Mid', 'acc_top': 'Top quarter'}
    psd_data = {}
    for ch_key, ch_name in channels.items():
        if data[ch_key] is None:
            continue
        t_fv, acc_fv, dt_fv = process_acceleration(
            data['t'], data[ch_key], t_fv_start)
        if acc_fv is None or len(acc_fv) < 100:
            continue
        fv_duration = float(t_fv[-1] - t_fv[0])
        win_sec = max(min(WELCH_WINDOW_SEC, fv_duration * 0.8), 2.0)
        freqs, psd_db = compute_psd(acc_fv, dt_fv, window_sec=win_sec)
        f1, _ = pick_peak(freqs, psd_db, MODE1_F_MIN, MODE1_F_MAX)
        result['{}_f1'.format(ch_key)] = f1
        result['{}_T1'.format(ch_key)] = 1.0 / f1 if f1 and f1 > 0 else None
        f2, _ = pick_peak(freqs, psd_db, MODE2_F_MIN, MODE2_F_MAX)
        result['{}_f2'.format(ch_key)] = f2
        psd_data[ch_key] = (freqs, psd_db, ch_name)

    wall_psds = []
    avg_freqs = None
    for ch_key in ['acc_bot', 'acc_mid', 'acc_top']:
        if ch_key in psd_data:
            freqs_ch, psd_ch, _ = psd_data[ch_key]
            wall_psds.append(10.0 ** (psd_ch / 10.0))
            avg_freqs = freqs_ch
    if wall_psds and avg_freqs is not None:
        avg_psd_linear = np.exp(np.mean(np.log(np.array(wall_psds) + 1e-30), axis=0))
        avg_psd_db = 10.0 * np.log10(avg_psd_linear + 1e-30)
        f1_avg, _ = pick_peak(avg_freqs, avg_psd_db, MODE1_F_MIN, MODE1_F_MAX)
        result['f1_best'] = f1_avg
        result['T1_best'] = 1.0 / f1_avg if f1_avg and f1_avg > 0 else None
        result['f1_source'] = 'avg_wall'
    else:
        for pref in ['acc_top', 'acc_mid', 'acc_bot']:
            f1 = result.get('{}_f1'.format(pref))
            if f1 is not None and f1 > 0:
                result['f1_best'] = f1
                result['T1_best'] = 1.0 / f1
                result['f1_source'] = pref
                break
        else:
            result['f1_best'] = None
            result['T1_best'] = None
            result['f1_source'] = None
    return result
