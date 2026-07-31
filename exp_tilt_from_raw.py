# -*- coding: ascii -*-
"""
Compute experimental wall tilt for Test 9 (US-1) from the raw shake-table
xlsx exports in EXP_DATA, using the same definitions as the 3DEC FISH
instruments (instrument_history_new.dat):

  tilt_bot_seg   = atan((d_bot  - d_table) / 0.66)   [deg]
  tilt_low_seg   = atan((d_mid  - d_bot)   / 0.60)
  tilt_up_seg    = atan((d_top  - d_mid)   / 0.80)
  tilt_full_wall = atan((d_top  - d_table) / 2.06)

Channel map (Test9_Info.xlsx): 1 = Disp Bot quarter (Z=0.66), 2 = Disp Mid
(Z=1.26), 3/4 = Disp Top quarter left/right (Z=2.06, Y=-/+0.54),
5 = Disp Shake table. Units m. IMPORTANT: the wall sensors are mounted on
the shake-table frame, so they already read wall displacement RELATIVE to
the table (verified: in Run 3 the table channel swings +-22 mm while the
wall channels move <0.2 mm). The table term is therefore NOT subtracted;
d_table is treated as zero. d_top is the mean of channels 3 and 4. All
channels are baselined to their Run 1 initial reading, so cumulative lean
is preserved across runs.

Metrics per run (matching postprocess_stratC.py):
  peak     = max |tilt(t) - tilt(t at run start)|   (re-zeroed per run)
  residual = mean of last 5% of tilt(t)             (cumulative, rel. Run 1)

Output: stratC_results_NODAMP_v6_NEW/postproc/exp_Test9_tilt.csv
Usage:  python exp_tilt_from_raw.py [EXP_DIR] [--runs LO HI]
"""
import argparse, csv, json, math, os
import numpy as np
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("pip install openpyxl")

HERE = Path(__file__).resolve().parent
DEF_EXP = r"C:\Users\yopi1\Documents\Itasca\3dec910\Hanze\Wall_Floor_Interaction\EXP_DATA"
STATE = Path(os.environ.get("EXP_TILT_STATE", "/tmp/exp_tilt_state.json"))

# Sensor elevations differ between specimens: the bot-quarter sensor is at
# Z=0.66 m in Test 9 (US-1) but Z=0.60 m in Test 12 (US-2, Test12_Info).
def segs_for(bot_z):
    return {  # name -> (num_hi, num_lo, dy)
        "tilt_bot_seg":   ("bot",   "table", bot_z),
        "tilt_low_seg":   ("mid",   "bot",   1.26 - bot_z),
        "tilt_up_seg":    ("top",   "mid",   0.80),
        "tilt_full_wall": ("top",   "table", 2.06),
    }
TEST_CFG = {9:  {"prefix": "Test9Run",  "bot_z": 0.66,
                 "out": "exp_Test9_tilt.csv"},
            12: {"prefix": "Test12Run", "bot_z": 0.60,
                 "out": "exp_Test12_tilt.csv"}}
TAIL_FRAC = 0.05
RAD2DEG = 180.0 / math.pi


def read_run(fpath):
    """Return dict of channel arrays (m) + time (s)."""
    wb = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header: (None,'Time',1..17)
    t, c1, c2, c3, c4, c5 = [], [], [], [], [], []
    for r in rows:
        if r[1] is None:
            continue
        t.append(r[1]); c1.append(r[2]); c2.append(r[3])
        c3.append(r[4]); c4.append(r[5]); c5.append(r[6])
    wb.close()
    a = lambda v: np.asarray(v, dtype=float)
    return {"t": a(t), "bot": a(c1), "mid": a(c2),
            "top": 0.5 * (a(c3) + a(c4)), "table": a(c5)}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("exp_dir", nargs="?", default=DEF_EXP)
    ap.add_argument("--runs", nargs=2, type=int, default=[1, 24])
    ap.add_argument("--test", type=int, default=9, choices=(9, 12))
    args = ap.parse_args()
    cfg = TEST_CFG[args.test]
    SEGS = segs_for(cfg["bot_z"])
    out_csv = (HERE / "stratC_results_NODAMP_v6_NEW" / "postproc"
               / cfg["out"])
    exp_dir = Path(args.exp_dir)
    lo, hi = args.runs

    # resumable state: run1 baselines + per-run metrics
    state = {"base": None, "runs": {}}
    if STATE.exists():
        state = json.loads(STATE.read_text())

    for rn in range(lo, hi + 1):
        f = exp_dir / ("%s%d.xlsx" % (cfg["prefix"], rn))
        if not f.exists():
            print("missing:", f.name); continue
        d = read_run(f)
        if state["base"] is None:
            if rn != 1:
                raise SystemExit("Run 1 must be processed first (baselines).")
            state["base"] = {k: float(d[k][0]) for k in
                             ("bot", "mid", "top", "table")}
        base = state["base"]
        rel = {k: d[k] - base[k] for k in ("bot", "mid", "top")}
        rel["table"] = np.zeros_like(d["bot"])  # sensors are table-mounted
        rec = {"n": int(len(d["t"])),
               "dur_s": round(float(d["t"][-1] - d["t"][0]), 2)}
        for name, (hi_k, lo_k, dy) in SEGS.items():
            tilt = np.arctan((rel[hi_k] - rel[lo_k]) / dy) * RAD2DEG
            tail = tilt[int((1.0 - TAIL_FRAC) * len(tilt)):]
            rec["peak_" + name] = round(
                float(np.max(np.abs(tilt - tilt[0]))), 5)
            rec["resid_" + name] = round(float(np.mean(tail)), 5)
        state["runs"][str(rn)] = rec
        print("run %2d: peak_full=%.4f deg  resid_full=%.4f deg  "
              "peak_up=%.4f deg" % (rn, rec["peak_tilt_full_wall"],
              rec["resid_tilt_full_wall"], rec["peak_tilt_up_seg"]))
        STATE.write_text(json.dumps(state))

    # write CSV of everything processed so far
    runs = sorted(int(k) for k in state["runs"])
    if runs:
        out_csv.parent.mkdir(parents=True, exist_ok=True)
        cols = (["run", "n", "dur_s"] +
                ["peak_" + s for s in SEGS] + ["resid_" + s for s in SEGS])
        with open(str(out_csv), "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(cols)
            for rn in runs:
                r = state["runs"][str(rn)]
                w.writerow([rn] + [r.get(c, "") for c in cols[1:]])
        print("-> {} ({} runs)".format(out_csv, len(runs)))


if __name__ == "__main__":
    main()
